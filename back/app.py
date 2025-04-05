# app.py
from flask import Flask, request, jsonify, send_file
import pyodbc
import os
from io import BytesIO
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from docxtpl import DocxTemplate
import json
from flask_cors import CORS

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Database configuration
DB_CONFIG = {
    'driver': '{ODBC Driver 17 for SQL Server}',
    'server': 'localhost',
    'database': 'AssetsManagement',
    'trusted_connection': 'yes'
    # For SQL authentication use:
    # 'uid': 'your_username',
    # 'pwd': 'your_password'
}

def get_db_connection():
    """Create a connection to the SQL Server database"""
    conn_str = (
        f"DRIVER={DB_CONFIG['driver']};"
        f"SERVER={DB_CONFIG['server']};"
        f"DATABASE={DB_CONFIG['database']};"
        f"Trusted_Connection={DB_CONFIG['trusted_connection']};"
    )
    # For SQL authentication
    # f"UID={DB_CONFIG['uid']};"
    # f"PWD={DB_CONFIG['pwd']};"
    return pyodbc.connect(conn_str)

# API endpoints

@app.route('/api/departments', methods=['GET'])
def get_departments():
    """Get all departments"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DepartmentID, DepartmentName FROM Departments")
        departments = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return jsonify(departments)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees', methods=['GET'])
def get_employees():
    """Get all employees with department information"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = """
        SELECT e.EmployeeID, e.IdentificationNumber, e.FullName, e.Position, d.DepartmentName, e.OfficeLocation 
        FROM Employees e
        JOIN Departments d ON e.DepartmentID = d.DepartmentID
        """
        cursor.execute(query)
        employees = []
        for row in cursor.fetchall():
            employees.append({
                'id': row[0],
                'identificationNumber': row[1],
                'fullName': row[2],
                'position': row[3],
                'department': row[4],
                'officeLocation': row[5]
            })
        cursor.close()
        conn.close()
        return jsonify(employees)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees/<int:employee_id>', methods=['GET'])
def get_employee(employee_id):
    """Get employee details by ID"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = """
        SELECT e.EmployeeID, e.IdentificationNumber, e.FullName, e.Position, e.DepartmentID, d.DepartmentName, e.OfficeLocation 
        FROM Employees e
        JOIN Departments d ON e.DepartmentID = d.DepartmentID
        WHERE e.EmployeeID = ?
        """
        cursor.execute(query, (employee_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Employee not found'}), 404
        
        employee = {
            'id': row[0],
            'identificationNumber': row[1],
            'fullName': row[2],
            'position': row[3],
            'departmentId': row[4],
            'department': row[5],
            'officeLocation': row[6]
        }
        cursor.close()
        conn.close()
        return jsonify(employee)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees', methods=['POST'])
def create_employee():
    """Create a new employee"""
    data = request.json
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = """
        INSERT INTO Employees (IdentificationNumber, FullName, Position, DepartmentID, OfficeLocation)
        VALUES (?, ?, ?, ?, ?)
        """
        cursor.execute(query, (
            data['identificationNumber'],
            data['fullName'],
            data['position'],
            data['departmentId'],
            data.get('officeLocation', '')
        ))
        conn.commit()
        new_id = cursor.execute("SELECT @@IDENTITY").fetchval()
        cursor.close()
        conn.close()
        return jsonify({'id': new_id, 'message': 'Employee created successfully'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/assets/categories', methods=['GET'])
def get_asset_categories():
    """Get all asset categories"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT CategoryID, CategoryName FROM AssetCategories")
        categories = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return jsonify(categories)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands', methods=['GET'])
def get_brands():
    """Get all brands"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT BrandID, BrandName FROM Brands")
        brands = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return jsonify(brands)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/assets', methods=['GET'])
def get_assets():
    """Get all assets with category and brand information"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = """
        SELECT a.AssetID, a.CategoryID, ac.CategoryName, a.BrandID, b.BrandName, a.Model, a.SerialNumber, a.Status
        FROM Assets a
        JOIN AssetCategories ac ON a.CategoryID = ac.CategoryID
        JOIN Brands b ON a.BrandID = b.BrandID
        """
        cursor.execute(query)
        assets = []
        for row in cursor.fetchall():
            assets.append({
                'id': row[0],
                'categoryId': row[1],
                'category': row[2],
                'brandId': row[3],
                'brand': row[4],
                'model': row[5],
                'serialNumber': row[6],
                'status': row[7]
            })
        cursor.close()
        conn.close()
        return jsonify(assets)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/assets', methods=['POST'])
def create_asset():
    """Create a new asset"""
    data = request.json
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = """
        INSERT INTO Assets (CategoryID, BrandID, Model, SerialNumber, Status)
        VALUES (?, ?, ?, ?, ?)
        """
        cursor.execute(query, (
            data['categoryId'],
            data['brandId'],
            data.get('model', ''),
            data.get('serialNumber', ''),
            data.get('status', 'Available')
        ))
        conn.commit()
        new_id = cursor.execute("SELECT @@IDENTITY").fetchval()
        cursor.close()
        conn.close()
        return jsonify({'id': new_id, 'message': 'Asset created successfully'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees/<int:employee_id>/assets', methods=['GET'])
def get_employee_assets(employee_id):
    """Get all assets assigned to an employee"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC GetEmployeeAssets ?", (employee_id,))
        assets = []
        for row in cursor.fetchall():
            assets.append({
                'assetId': row[0],
                'category': row[1],
                'brand': row[2],
                'model': row[3],
                'serialNumber': row[4],
                'assignmentDate': row[5].strftime('%Y-%m-%d') if row[5] else None,
                'observations': row[6]
            })
        cursor.close()
        conn.close()
        return jsonify(assets)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees/<int:employee_id>/software', methods=['GET'])
def get_employee_software(employee_id):
    """Get all software licenses assigned to an employee"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC GetEmployeeSoftware ?", (employee_id,))
        software = []
        for row in cursor.fetchall():
            software.append({
                'licenseName': row[0],
                'username': row[1],
                'password': '********',  # Mask the password
                'requiresPasswordChange': bool(row[3])
            })
        cursor.close()
        conn.close()
        return jsonify(software)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees/<int:employee_id>/credentials', methods=['GET'])
def get_employee_credentials(employee_id):
    """Get all special credentials assigned to an employee"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC GetEmployeeCredentials ?", (employee_id,))
        credentials = []
        for row in cursor.fetchall():
            credentials.append({
                'url': row[0],
                'username': row[1],
                'password': '********',  # Mask the password
                'requiresPasswordChange': bool(row[3])
            })
        cursor.close()
        conn.close()
        return jsonify(credentials)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/assign/asset', methods=['POST'])
def assign_asset():
    """Assign an asset to an employee"""
    data = request.json
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if asset is available
        cursor.execute("SELECT Status FROM Assets WHERE AssetID = ?", (data['assetId'],))
        status = cursor.fetchone()[0]
        if status != 'Available':
            return jsonify({'error': 'Asset is not available for assignment'}), 400
        
        # Update asset status
        cursor.execute("UPDATE Assets SET Status = 'Assigned' WHERE AssetID = ?", (data['assetId'],))
        
        # Create assignment
        query = """
        INSERT INTO AssetAssignments (EmployeeID, AssetID, AssignmentDate, Observations)
        VALUES (?, ?, ?, ?)
        """
        cursor.execute(query, (
            data['employeeId'],
            data['assetId'],
            data.get('assignmentDate', datetime.now().strftime('%Y-%m-%d')),
            data.get('observations', '')
        ))
        
        conn.commit()
        new_id = cursor.execute("SELECT @@IDENTITY").fetchval()
        cursor.close()
        conn.close()
        return jsonify({'id': new_id, 'message': 'Asset assigned successfully'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel/<int:employee_id>', methods=['GET'])
def export_excel(employee_id):
    """Export employee assets to Excel"""
    try:
        # Get employee data
        employee_data = get_employee_data(employee_id)
        
        if not employee_data or len(employee_data) < 1 or len(employee_data[0]) < 1:
            return jsonify({'error': 'No employee data found'}), 404
        
        # Create Excel workbook
        workbook = create_excel_report(employee_data)
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Get employee name for filename
        employee_name = employee_data[0][0]['fullName'].replace(' ', '_')
        
        # Return the Excel file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Assets_{employee_name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/word/<int:employee_id>', methods=['GET'])
def export_word(employee_id):
    """Generate and export Word document (Acta)"""
    try:
        # Get employee data
        employee_data = get_employee_data(employee_id)
        
        if not employee_data or len(employee_data) < 1 or len(employee_data[0]) < 1:
            return jsonify({'error': 'No employee data found'}), 404
        
        # Generate Word document
        doc_buffer = generate_word_document(employee_data)
        doc_buffer.seek(0)
        
        # Get employee name for filename
        employee_name = employee_data[0][0]['fullName'].replace(' ', '_')
        
        # Return the Word document
        return send_file(
            doc_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            as_attachment=True,
            download_name=f'Acta_Entrega_{employee_name}_{datetime.now().strftime("%Y%m%d")}.docx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Helper functions

def get_employee_data(employee_id):
    """Get all data needed for reports for a specific employee"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Employee info
        employee_query = """
        SELECT 
            e.FullName,
            e.IdentificationNumber,
            e.Position,
            d.DepartmentName,
            e.OfficeLocation,
            FORMAT(GETDATE(), 'dd/MM/yyyy') AS CurrentDate
        FROM Employees e
        JOIN Departments d ON e.DepartmentID = d.DepartmentID
        WHERE e.EmployeeID = ?
        """
        cursor.execute(employee_query, (employee_id,))
        employee_rows = cursor.fetchall()
        employee_info = []
        for row in employee_rows:
            employee_info.append({
                'fullName': row[0],
                'identificationNumber': row[1],
                'position': row[2],
                'department': row[3],
                'officeLocation': row[4],
                'currentDate': row[5]
            })
        
        # Hardware assets
        cursor.execute("EXEC GetEmployeeAssets ?", (employee_id,))
        assets = []
        for row in cursor.fetchall():
            assets.append({
                'assetId': row[0],
                'category': row[1],
                'brand': row[2],
                'model': row[3],
                'serialNumber': row[4],
                'assignmentDate': row[5].strftime('%Y-%m-%d') if row[5] else None,
                'observations': row[6] if row[6] else ''
            })
        
        # Software licenses
        cursor.execute("EXEC GetEmployeeSoftware ?", (employee_id,))
        software = []
        for row in cursor.fetchall():
            software.append({
                'licenseName': row[0],
                'username': row[1],
                'password': '********',  # Mask the password
                'requiresPasswordChange': bool(row[3])
            })
        
        # Special credentials
        cursor.execute("EXEC GetEmployeeCredentials ?", (employee_id,))
        credentials = []
        for row in cursor.fetchall():
            credentials.append({
                'url': row[0],
                'username': row[1],
                'password': '********',  # Mask the password
                'requiresPasswordChange': bool(row[3])
            })
        
        cursor.close()
        conn.close()
        
        return [employee_info, assets, software, credentials]
    except Exception as e:
        print(f"Error getting employee data: {str(e)}")
        return None

def create_excel_report(employee_data):
    """Create an Excel report with employee data"""
    employee_info, assets, software, credentials = employee_data
    employee = employee_info[0]
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Employee Info Sheet
    ws_info = wb.active
    ws_info.title = "Información Empleado"
    
    # Add employee info
    ws_info['A1'] = "Nombre"
    ws_info['B1'] = employee['fullName']
    ws_info['A2'] = "Identificación"
    ws_info['B2'] = employee['identificationNumber']
    ws_info['A3'] = "Cargo"
    ws_info['B3'] = employee['position']
    ws_info['A4'] = "Dependencia"
    ws_info['B4'] = employee['department']
    ws_info['A5'] = "Ubicación"
    ws_info['B5'] = employee['officeLocation']
    ws_info['A6'] = "Fecha"
    ws_info['B6'] = employee['currentDate']
    
    # Style the info cells
    for row in range(1, 7):
        cell = ws_info.cell(row=row, column=1)
        cell.font = header_font
        cell.border = border
    
    # Hardware Assets Sheet
    if assets:
        ws_assets = wb.create_sheet("Hardware")
        
        # Add headers
        headers = ["Categoría", "Marca", "Modelo", "Serial", "Fecha Asignación", "Observaciones"]
        for col, header in enumerate(headers, start=1):
            cell = ws_assets.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Add asset data
        for row, asset in enumerate(assets, start=2):
            ws_assets.cell(row=row, column=1).value = asset['category']
            ws_assets.cell(row=row, column=2).value = asset['brand']
            ws_assets.cell(row=row, column=3).value = asset['model']
            ws_assets.cell(row=row, column=4).value = asset['serialNumber']
            ws_assets.cell(row=row, column=5).value = asset['assignmentDate']
            ws_assets.cell(row=row, column=6).value = asset['observations']
            
            # Add borders
            for col in range(1, len(headers) + 1):
                ws_assets.cell(row=row, column=col).border = border
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws_assets.column_dimensions[chr(64 + col)].width = 15
    
    # Software Licenses Sheet
    if software:
        ws_software = wb.create_sheet("Software")
        
        # Add headers
        headers = ["Licencia", "Usuario", "Requiere Cambio Contraseña"]
        for col, header in enumerate(headers, start=1):
            cell = ws_software.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Add software data
        for row, sw in enumerate(software, start=2):
            ws_software.cell(row=row, column=1).value = sw['licenseName']
            ws_software.cell(row=row, column=2).value = sw['username']
            ws_software.cell(row=row, column=3).value = "Sí" if sw['requiresPasswordChange'] else "No"
            
            # Add borders
            for col in range(1, len(headers) + 1):
                ws_software.cell(row=row, column=col).border = border
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws_software.column_dimensions[chr(64 + col)].width = 20
    
    # Special Credentials Sheet
    if credentials:
        ws_creds = wb.create_sheet("Credenciales")
        
        # Add headers
        headers = ["URL", "Usuario", "Requiere Cambio Contraseña"]
        for col, header in enumerate(headers, start=1):
            cell = ws_creds.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Add credential data
        for row, cred in enumerate(credentials, start=2):
            ws_creds.cell(row=row, column=1).value = cred['url']
            ws_creds.cell(row=row, column=2).value = cred['username']
            ws_creds.cell(row=row, column=3).value = "Sí" if cred['requiresPasswordChange'] else "No"
            
            # Add borders
            for col in range(1, len(headers) + 1):
                ws_creds.cell(row=row, column=col).border = border
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws_creds.column_dimensions[chr(64 + col)].width = 25
    
    return wb

def generate_word_document(employee_data):
    """Generate Word document using a template"""
    employee_info, assets, software, credentials = employee_data
    employee = employee_info[0]
    
    # Load the template
    template_path = os.path.join(os.path.dirname(__file__), 'templates', 'Acta-Entrega-Recursos-Tics-Plantila.docx')
    doc = DocxTemplate(template_path)
    
    # Create context with all data
    context = {
        'FECHA': employee['currentDate'],
        'NOMBRE_EMPLEADO': employee['fullName'],
        'IDENTIFICACION': employee['identificationNumber'],
        'CARGO': employee['position'],
        'DEPENDENCIA': employee['department'],
        'UBICACION': employee['officeLocation'],
        'hardware': [],
        'software': [],
        'credentials': []
    }
    
    # Organize hardware by category
    hardware_by_category = {}
    for asset in assets:
        category = asset['category']
        if category not in hardware_by_category:
            hardware_by_category[category] = []
        hardware_by_category[category].append(asset)
    
    # Format hardware for the template
    for category, items in hardware_by_category.items():
        context['hardware'].append({
            'categoria': category,
            'cantidad': len(items),
            'marca': items[0]['brand'] if items else '',
            'modelo': items[0]['model'] if items else '',
            'serial': items[0]['serialNumber'] if items else '',
            'observaciones': items[0]['observations'] if items else ''
        })
    
    # Format software for the template
    for sw in software:
        context['software'].append({
            'nombre': sw['licenseName'],
            'usuario': sw['username'],
            'password': '********',
            'cambio': 'Sí' if sw['requiresPasswordChange'] else 'No'
        })
    
    # Format credentials for the template
    for cred in credentials:
        context['credentials'].append({
            'url': cred['url'],
            'usuario': cred['username'],
            'password': '********',
            'cambio': 'Sí' if cred['requiresPasswordChange'] else 'No'
        })
    
    # Render the template
    doc.render(context)
    
    # Save to BytesIO
    output = BytesIO()
    doc.save(output)
    output.seek(0)
    
    return output

if __name__ == '__main__':
    app.run(debug=True)